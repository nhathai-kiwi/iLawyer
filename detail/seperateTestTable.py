#!/usr/bin/env python
# -*- coding: utf-8 -*- 
#!/usr/bin/python3
import sys  
reload(sys)  
sys.setdefaultencoding('utf8')
import json
import pprint
#import mysql.connector
#from mysql.connector import (connection)
import re
import codecs
import myTokenizer as mt
import openpyxl

def processText(ith, mainText):
	ret = ""
	ret += "Câu hỏi " + str(int(ith)) + ":\n"
	x = mt.MyTokenizer(mainText)
	words = x.printOut()
	for i in words:
		ret += "\t" + i + "\n"
	return ret

def readFileXlsx():
	book = openpyxl.load_workbook('testTableTraining.xlsx')
	sheet = book.active
	f = open('vnTokenizer.txt', "w")
	for row in range(2, 102):
		ith = sheet.cell(row = row, column = 1).value
		mainText = sheet.cell(row = row, column = 2).value
		ret = processText(ith, mainText)
		f.write(ret)
readFileXlsx()

def printFileXlsx():
	book = openpyxl.load_workbook('testTableTraining.xlsx')
	sheet = book.active
	f = open('question.txt', "w")
	for row in range(2, 102):
		#ith = sheet.cell(row = row, column = 1).value
		mainText = sheet.cell(row = row, column = 2).value
		#ret = processText(ith, mainText)
		#f.write(ret)
		f.write("Question " + str(row  - 1) + ":\n")
		f.write(mainText + "\n")
printFileXlsx()